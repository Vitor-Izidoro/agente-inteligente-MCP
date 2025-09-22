import speech_recognition as sr
import pyttsx3
import openpyxl
import requests
import langchain
import pyaudio
# tamo usando llama3
import ollama
import requests
import webbrowser
import re  # n sei oq essa bib faz
import json
from typing import Optional, Tuple, Any

# Função para criar/abrir planilha de produtos
def carregar_planilha(nome_arquivo="lista_compras.xlsx"):
    try:
        wb = openpyxl.load_workbook(nome_arquivo)
        ws = wb["Produtos"]
        # Garante que as colunas existam
        if ws.max_row == 0 or ws.cell(row=1, column=1).value != "Produto":
            ws.append(["Produto", "Preço"])
            wb.save(nome_arquivo)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produtos"
        ws.append(["Produto", "Preço"])
        wb.save(nome_arquivo)
    return wb

# Função para listar produtos e preços
def listar_produtos(nome_arquivo="lista_compras.xlsx"):
    wb = carregar_planilha(nome_arquivo)
    ws = wb["Produtos"]
    produtos = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            produtos.append(f"{row[0]} - R$ {row[1] if row[1] else 'N/A'}")
    return produtos

# Função para remover produto (por nome)
def remover_produto(produto, nome_arquivo="lista_compras.xlsx"):
    wb = carregar_planilha(nome_arquivo)
    ws = wb["Produtos"]
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] and produto.lower() == row[0].lower():
            ws.delete_rows(idx)
            wb.save(nome_arquivo)
            break

# Função para gerar link WhatsApp (mostra nome e preço)
def gerar_link_whatsapp(produtos):
    texto = "Lista de compras: " + ", ".join(produtos)
    numero = "5541988039241"
    link = f"https://wa.me/{numero}?text={requests.utils.quote(texto)}"
    webbrowser.open(link)
    return link

# Função para adicionar produto com preço opcional
def adicionar_produto(produto, preco=None, nome_arquivo="lista_compras.xlsx"):
    wb = carregar_planilha(nome_arquivo)
    ws = wb["Produtos"]
    if preco is not None:
        ws.append([produto, preco])
    else:
        ws.append([produto, ""])
    wb.save(nome_arquivo)

# ================================#
# Interpretação via LLM (Ollama)  #
# ================================#

LLM_MODEL = "llama3"  # modelo usado LOCAL

LLM_SYSTEM_PROMPT = (
    "Você é um parser de comandos de lista de compras. \n"
    "Recebe uma frase em português informal e DEVE responder somente um JSON válido, sem explicações. \n"
    "Campos: action (adicionar|remover|listar|enviar|sair|desconhecido), product (string ou null), price (float ou null). \n"
    "Regras: \n- Se ação envolver adicionar ou remover, tente extrair o produto. \n"
    "- Produto é o nome livre após remover verbos de ação. \n"
    "- Preço: detectar número (inteiro ou decimal) possivelmente seguido de 'reais', 'real', 'R$', 'rs'. Converter vírgula para ponto. \n"
    "- Se não houver preço, usar null. \n"
    "- Se ação não reconhecida: action=desconhecido. \n"
    "Responda somente JSON."
)

def chamar_llm(frase: str) -> Optional[dict[str, Any]]:
    try:
        resp = ollama.chat(model=LLM_MODEL, messages=[
            {"role": "system", "content": LLM_SYSTEM_PROMPT},
            {"role": "user", "content": frase}
        ])
        conteudo = resp.get("message", {}).get("content", "").strip()
        # Tentar isolar JSON (caso o modelo responda com texto extra)
        inicio = conteudo.find('{')
        fim = conteudo.rfind('}')
        if inicio != -1 and fim != -1:
            conteudo = conteudo[inicio:fim+1]
        return json.loads(conteudo)
    except Exception as e:
        print(f"[LLM] Falha ao interpretar via LLM: {e}")
        return None

def interpretar_comando_llm(frase: str) -> Tuple[Optional[str], Optional[str], Optional[float]]:
    if not frase:
        return None, None, None
    data = chamar_llm(frase)
    if not data:
        return None, None, None
    action = data.get("action")
    product = data.get("product") or None
    price = data.get("price")
    # Normalizações
    if isinstance(action, str):
        action = action.lower().strip()
    if action not in {"adicionar", "remover", "listar", "enviar", "sair", "desconhecido"}:
        action = "desconhecido"
    if isinstance(price, str):
        try:
            price = float(price.replace(',', '.'))
        except Exception:
            price = None
    if isinstance(price, (int, float)):
        price = float(price)
    else:
        price = None if price is not None and price != price else price  # trata NaN
    return (action if action != "desconhecido" else None), product, price

# Função para reconhecer comando de voz
def reconhecer_comando():
    r = sr.Recognizer()
    with sr.Microphone() as source:
        print("Fale seu comando...")
        audio = r.listen(source)
    try:
        comando = r.recognize_google(audio, language="pt-BR")
        print(f"Você disse: {comando}")
        return comando.lower()
    except sr.UnknownValueError:
        print("Não entendi o comando.")
        return ""
    except sr.RequestError:
        print("Erro ao acessar o serviço de reconhecimento.")
        return ""

# Função para reconhecer palavra de ativação

def ouvir_ate_ativacao(wake_word="oi"):
    r = sr.Recognizer()
    while True:
        with sr.Microphone() as source:
            print(f"Diga '{wake_word}' para ativar o agente...")
            audio = r.listen(source)
        try:
            texto = r.recognize_google(audio, language="pt-BR")
            print(f"Você disse: {texto}")
            if wake_word in texto.lower():
                print("oi ativada! Fale seu comando...")
                return True
        except sr.UnknownValueError:
            print("Não entendi. Tente novamente.")
        except sr.RequestError:
            print("Erro ao acessar o serviço de reconhecimento.")

# Função principal 
def main():
    print("Agente de lista de compras iniciado (interpretação via LLM apenas).")
    while True:
        ouvir_ate_ativacao()
        frase = reconhecer_comando()
        acao, produto, preco = interpretar_comando_llm(frase)
        if not acao:
            print("Não entendi a ação. Tente novamente.")
            continue
        if acao == "sair":
            print("Encerrando agente.")
            break
        if acao == "listar":
            produtos = listar_produtos()
            print("Produtos:", produtos if produtos else "(lista vazia)")
            continue
        if acao == "enviar":
            produtos = listar_produtos()
            link = gerar_link_whatsapp(produtos)
            print("Link para WhatsApp:", link)
            continue
        if acao == "adicionar":
            if not produto:
                print("Não identifiquei o nome do produto para adicionar.")
                continue
            adicionar_produto(produto, preco)
            if preco is not None:
                print(f"Produto '{produto}' adicionado com preço R$ {preco}.")
            else:
                print(f"Produto '{produto}' adicionado sem preço.")
            continue
        if acao == "remover":
            if not produto:
                print("Não identifiquei o nome do produto para remover.")
                continue
            remover_produto(produto)
            print(f"Produto '{produto}' removido.")
            continue

if __name__ == "__main__":
    main()